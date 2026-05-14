#!/usr/bin/env python3
from __future__ import annotations
import os
import re
import shutil
import sys
import random
import urllib.request
import subprocess
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import List, Dict, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

IMG_EXT = {"png", "jpg", "jpeg", "webp", "heic"}
XLS_EXT = {"xlsx", "xls"}
REQ_HEADERS = {"子任务ID", "总页数", "总评测数量"}


@dataclass
class TaskRow:
    order: int
    cols: List
    subtask_id: str
    pages: float
    tags: float


def split_files(raw: str) -> list[Path]:
    out = []
    seen = set()
    for part in raw.split("|"):
        p = Path(part.strip())
        if not part.strip() or part in seen:
            continue
        seen.add(part)
        if p.exists() and p.is_file():
            out.append(p)
    return out


def has_required_headers(path: Path) -> bool:
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        for r in range(1, min(20, ws.max_row) + 1):
            vals = {str(ws.cell(r, c).value or "").strip() for c in range(3, 14)}
            if REQ_HEADERS.issubset(vals):
                return True
    except Exception:
        return False
    return False


def classify(files: list[Path]):
    shots = []
    ai = None
    card = None
    others = []
    for f in files:
        ext = f.suffix.lower().lstrip(".")
        name = f.name
        if ext in IMG_EXT:
            shots.append(f)
        elif ext in XLS_EXT:
            if ai is None and "AI_待分配" in name:
                ai = f
            elif card is None and "答题卡_待分配" in name:
                card = f
            else:
                others.append(f)

    valid_others = [x for x in others if has_required_headers(x)]
    if ai is None and valid_others:
        ai = valid_others.pop(0)
    if card is None and valid_others:
        card = valid_others.pop(0)
    return shots, ai, card


def mock_download(ai_target: Path, card_target: Path) -> bool:
    ai_src = os.environ.get("DAILY_ASSIGN_MOCK_AI_SOURCE", "").strip()
    card_src = os.environ.get("DAILY_ASSIGN_MOCK_CARD_SOURCE", "").strip()
    if not ai_src or not card_src:
        return False
    ai_s = Path(ai_src)
    card_s = Path(card_src)
    if not ai_s.exists() or not card_s.exists():
        return False
    shutil.copy2(ai_s, ai_target)
    shutil.copy2(card_s, card_target)
    print("[下载] 已覆盖旧的同名文件")
    return True


def download_with_retry(url: str, target: Path) -> bool:
    if not url:
        return False
    for i in range(3):
        try:
            with urllib.request.urlopen(url, timeout=30) as r:
                data = r.read()
            target.write_bytes(data)
            return True
        except Exception as e:
            if i == 2:
                print(f"[下载] 失败: {target.name} ({e})")
            else:
                print(f"[下载] 重试 {i+1}/2: {target.name}")
    return False


def real_download(ai_target: Path, card_target: Path) -> bool:
    ai_url = os.environ.get("DAILY_ASSIGN_AI_URL", "").strip()
    card_url = os.environ.get("DAILY_ASSIGN_CARD_URL", "").strip()
    ok_ai = download_with_retry(ai_url, ai_target)
    ok_card = download_with_retry(card_url, card_target)
    if ok_ai and ok_card:
        print("[下载] 已覆盖旧的同名文件")
        return True
    return False


def to_num(v) -> float:
    try:
        if v is None:
            return 0.0
        return float(v)
    except Exception:
        return 0.0


def read_tasks(path: Path, cap: int) -> Tuple[List[str], List[TaskRow]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row = None
    header_idx = {}
    source_headers = []
    for r in range(1, min(20, ws.max_row) + 1):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(3, 14)]
        if REQ_HEADERS.issubset(set(vals)):
            header_row = r
            source_headers = vals
            for i, v in enumerate(vals, start=3):
                if v:
                    header_idx[v] = i
            break
    if header_row is None:
        raise ValueError("E003：今天任务的表格 格式不符合要求")

    out: List[TaskRow] = []
    order = 0
    for r in range(header_row + 1, ws.max_row + 1):
        subtask = str(ws.cell(r, header_idx["子任务ID"]).value or "").strip()
        pages = to_num(ws.cell(r, header_idx["总页数"]).value)
        tags = to_num(ws.cell(r, header_idx["总评测数量"]).value)
        if not subtask or pages <= 0:
            continue
        cols = [ws.cell(r, c).value for c in range(3, 14)]
        out.append(TaskRow(order=order, cols=cols, subtask_id=subtask, pages=pages, tags=tags))
        order += 1

    rnd = random.Random(int(date.today().strftime("%Y%m%d")))
    rnd.shuffle(out)

    picked: List[TaskRow] = []
    s = 0.0
    for row in out:
        if not picked and row.pages > cap:
            picked.append(row)
            break
        if s + row.pages <= cap:
            picked.append(row)
            s += row.pages
        if cap > 0 and picked and (cap - s) / cap < 0.05:
            break
    return source_headers, picked


def lev(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    dp = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        prev = dp[0]
        dp[0] = i
        for j, cb in enumerate(b, 1):
            cur = dp[j]
            dp[j] = min(dp[j] + 1, dp[j - 1] + 1, prev + (0 if ca == cb else 1))
            prev = cur
    return dp[-1]


def normalize_name(raw: str, names: List[str]) -> str | None:
    if raw in names:
        return raw
    cands = [(lev(raw, n), n) for n in names]
    cands.sort(key=lambda x: x[0])
    if cands and cands[0][0] <= 1:
        if len(cands) == 1 or cands[1][0] > cands[0][0]:
            return cands[0][1]
    return None


def run_vision_ocr(image_path: Path) -> str:
    script = Path(__file__).with_name("ocr_vision.swift")
    if not script.exists():
        return ""
    try:
        p = subprocess.run(["/usr/bin/swift", str(script), str(image_path)], capture_output=True, text=True, timeout=30)
        if p.returncode == 0:
            return p.stdout
    except Exception:
        return ""
    return ""


def parse_signup_from_shots(shots: List[Path], names: List[str]) -> Tuple[Dict[str, int], List[str]]:
    out: Dict[str, int] = {}
    unmatched: List[str] = []
    for shot in shots:
        text_file = shot.with_suffix(".txt")
        parsed_any = False
        text = ""
        if text_file.exists():
            text = text_file.read_text(encoding="utf-8", errors="ignore")
        else:
            text = run_vision_ocr(shot)

        for raw_name, cnt_s in re.findall(r"([\u4e00-\u9fa5]{2,4})\D{0,8}(\d+)", text):
            parsed_any = True
            norm = normalize_name(raw_name, names)
            if not norm:
                unmatched.append(raw_name)
                continue
            cnt = int(cnt_s)
            if cnt > 0:
                out[norm] = cnt

        if parsed_any:
            continue

        stem = shot.stem.replace("（", "(").replace("）", ")")
        m = re.search(r"([\u4e00-\u9fa5]{2,4})\D*(\d+)", stem)
        if m:
            raw_name, cnt_s = m.group(1), m.group(2)
            norm = normalize_name(raw_name, names)
            if not norm:
                unmatched.append(raw_name)
                continue
            cnt = int(cnt_s)
            if cnt > 0:
                out[norm] = cnt
    return out, unmatched


def assign(tasks: List[TaskRow], signup: Dict[str, int], weight_key: str, carry: Dict[str, float] | None = None):
    names = sorted(signup.keys())
    total_signup = sum(signup.values())
    target = {n: signup[n] / total_signup for n in names}
    assigned_weight = {n: 0.0 for n in names}
    assigned_count = {n: 0 for n in names}
    if carry:
        for n in names:
            assigned_weight[n] += carry.get(n, 0.0)

    result = []
    total_w = 0.0
    for t in tasks:
        w = t.pages if weight_key == "page" else t.tags
        if w <= 0:
            w = 1.0
        total_w += w
        best = None
        best_key = None
        for n in names:
            debt = target[n] * max(total_w, 1.0) - assigned_weight[n]
            key = (debt, -assigned_count[n], n)
            if best is None or key > best_key:
                best = n
                best_key = key
        assigned_weight[best] += w
        assigned_count[best] += 1
        result.append((t, best))

    grouped = {n: [] for n in names}
    for item in result:
        grouped[item[1]].append(item)
    final = []
    for n in names:
        final.extend(grouped[n])
    return final, assigned_weight


def delivery_text(today: date) -> str:
    d = today + timedelta(days=1)
    text = f"{d.month}月{d.day}日"
    if d.weekday() == 3:
        return text + " 14:00"
    return text


def period_text(today: date) -> str:
    if today.day >= 25:
        start = today.replace(day=25)
        if today.month == 12:
            end = date(today.year + 1, 1, 24)
        else:
            end = date(today.year, today.month + 1, 24)
    else:
        if today.month == 1:
            start = date(today.year - 1, 12, 25)
        else:
            start = date(today.year, today.month - 1, 25)
        end = today.replace(day=24)
    return f"{start.month:02d}{start.day:02d}-{end.month:02d}{end.day:02d}"


def write_sheet(ws, source_headers: List[str], rows, today: date):
    headers = ["周期"] + source_headers + ["负责人", "交付日期", "", "", "", "", "", ""]
    ws.append(headers)
    yellow = PatternFill(fill_type="solid", fgColor="FFF4B084")
    ws["E1"].fill = yellow
    ws["J1"].fill = yellow
    ws["K1"].fill = yellow
    ws.auto_filter.ref = "A1:T1"

    period = period_text(today)
    delivery = delivery_text(today)
    for task, owner in rows:
        row = [period] + task.cols + [owner, delivery, "", "", "", "", "", ""]
        ws.append(row)


def ai_task_name(today: date) -> str:
    y = today - timedelta(days=1)
    return f"{y.strftime('%y%m%d')}-{today.strftime('%y%m%d')}"


def card_task_name(today: date) -> str:
    wd = today.weekday()
    offset_to_wed = (wd - 2) % 7
    start = today - timedelta(days=offset_to_wed)
    end = start + timedelta(days=6)
    return f"{start.strftime('%m%d')}-{end.strftime('%m%d')}答题卡周期评测"


def main() -> int:
    files = split_files(os.environ.get("DAILY_ASSIGN_FILES", ""))
    shots, ai, card = classify(files)

    if not shots:
        print("E001：未检测到 有效 报名截图")
        return 1

    dl_dir = Path(os.environ.get("DOWNLOAD_DIR", str(Path.home() / "Downloads")))
    dl_dir.mkdir(parents=True, exist_ok=True)

    if ai is None and card is None:
        print("[下载] 检测到仅上传截图，开始自动下载今日任务表...")
        print(f"[下载] AI任务名: {ai_task_name(date.today())}")
        print(f"[下载] 答题卡任务名: {card_task_name(date.today())}")
        mode = os.environ.get("DAILY_ASSIGN_DOWNLOAD_MODE", "disabled").strip().lower()
        ai_path = dl_dir / "AI_待分配.xlsx"
        card_path = dl_dir / "答题卡_待分配.xlsx"
        ok = False
        if mode == "mock":
            ok = mock_download(ai_path, card_path)
        elif mode == "real":
            ok = real_download(ai_path, card_path)
        if not ok:
            print("E002：自动下载失败 且 缺少 今天任务的表格")
            print("请手动上传源文件继续")
            return 1
        ai, card = ai_path, card_path

    names_env = os.environ.get("NAMES", "").strip()
    names = [x.strip() for x in names_env.split(",") if x.strip()]
    signup, unmatched = parse_signup_from_shots(shots, names)
    if sum(signup.values()) <= 0:
        print("E004：OCR识别 - 无有效报名人数量")
        return 1

    method = os.environ.get("DAILY_ASSIGN_METHOD", "page")
    mode = os.environ.get("DAILY_ASSIGN_MODE", "linked")
    ai_cap = int(os.environ.get("DAILY_ASSIGN_AI_MAX", "200"))
    card_cap = int(os.environ.get("DAILY_ASSIGN_CARD_MAX", "300"))

    ai_rows = []
    card_rows = []
    ai_headers = ["任务名称", "子任务顺序", "任务ID", "子任务ID", "线上学生作业ID", "老师作业ID", "题单ID", "未评测页数", "总页数", "总评测数量", "任务链接"]
    card_headers = ai_headers[:]
    carry = None
    if ai is not None:
        ai_headers, ai_tasks = read_tasks(ai, ai_cap)
        ai_rows, carry = assign(ai_tasks, signup, method)
    if card is not None:
        card_headers, card_tasks = read_tasks(card, card_cap)
        card_rows, _ = assign(card_tasks, signup, method, carry if mode == "linked" else None)

    output_dir = Path(os.environ.get("OUTPUT_DIR", str(Path.cwd())))
    output = output_dir / "分配表.xlsx"
    tmp = output.with_suffix(".tmp.xlsx")

    try:
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        ws_ai = wb.create_sheet("AI")
        write_sheet(ws_ai, ai_headers, ai_rows, date.today())

        ws_card = wb.create_sheet("答题卡")
        write_sheet(ws_card, card_headers, card_rows, date.today())

        wb.save(tmp)
        os.replace(tmp, output)
    except Exception as e:
        print("E005：写出 分配表 失败")
        print(f"错误：{e}")
        return 1

    print("[分配] 报名人+报名数量/比例+实际分配数量/比例：")
    total = sum(signup.values())
    assigned = {}
    for t, owner in ai_rows + card_rows:
        assigned[owner] = assigned.get(owner, 0) + 1
    all_names = sorted(set(names))
    total_tasks = len(ai_rows + card_rows)
    for n in all_names:
        signup_n = signup.get(n, 0)
        assign_n = assigned.get(n, 0)
        signup_ratio = (signup_n / total * 100.0) if total > 0 else 0.0
        assign_ratio = (assign_n / max(total_tasks, 1) * 100.0)
        warn = " ⚠️" if signup_n == 0 or assign_n == 0 else ""
        print(f"- {n}: 报名 {signup_n}/{signup_ratio:.1f}% | 分配 {assign_n}/{assign_ratio:.1f}%{warn}")
    for bad in sorted(set(unmatched)):
        print(f"- {bad}: 报名 无法匹配 | 分配 0/0.0% ⚠️")

    print("已生成：分配表.xlsx 📁")
    print("👉 任务已完成")
    return 0


if __name__ == "__main__":
    sys.exit(main())
