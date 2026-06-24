#!/usr/bin/env python3
from __future__ import annotations
import os
import re
import shutil
import sys
import json
import subprocess
import warnings
import time
import random
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import List, Dict, Tuple

# Suppress openpyxl UserWarning
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

if __name__ == "__main__" and not os.environ.get("DAILY_ASSIGN_FILES", "").strip():
    print("E001：未检测到 有效 报名截图")
    sys.exit(1)

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    sync_playwright = None

IMG_EXT = {"png", "jpg", "jpeg", "webp", "heic"}
XLS_EXT = {"xlsx", "xls"}
REQ_HEADERS = {"子任务ID", "总页数", "总评测数量"}


@dataclass
class TaskRow:
    order: int
    cols: List
    subtask_id: str
    pages: float
    tags: float


@dataclass(frozen=True)
class CheckGroup:
    sheet: str
    key: str
    subtask_id: str
    row_indices: tuple[int, ...]
    owners: tuple[str, ...]
    pages: float
    first_row_index: int


CHECKERS = ["郭小雨", "符于娜", "李橙橙"]


def split_files(raw: str) -> list[Path]:
    out = []
    seen = set()
    for part in raw.split("|"):
        p = Path(part.strip())
        if not part.strip() or part in seen:
            continue
        seen.add(part)
        if p.exists() and p.is_file():
            out.append(p)
    return out


def has_required_headers(path: Path) -> bool:
    try:
        if path.stat().st_size < 1000: return False
        wb = load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        for r in range(1, min(20, ws.max_row) + 1):
            for c_start in range(1, 10):
                vals = {str(ws.cell(r, c).value or "").strip() for c in range(c_start, c_start + 11)}
                if REQ_HEADERS.issubset(vals):
                    return True
    except Exception:
        return False
    return False


def task_table_kind(path: Path) -> str | None:
    try:
        if path.stat().st_size < 1000:
            return None
        wb = load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        max_row = min(ws.max_row or 0, 80)
        max_col = min(ws.max_column or 0, 20)
        texts: List[str] = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            for value in row:
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    texts.append(text)
        if not texts:
            return None
        joined = "\n".join(texts)
        lower = joined.lower()
        card_score = 0
        ai_score = 0
        if "cardpage" in lower:
            card_score += 5
        if "答题卡" in joined:
            card_score += 4
        if "holepage" in lower:
            ai_score += 5
        if "(auto)" in lower:
            ai_score += 3
        if "ai混合" in lower or "分数&ai" in lower:
            ai_score += 3
        if card_score > ai_score and card_score > 0:
            return "card"
        if ai_score > card_score and ai_score > 0:
            return "ai"
    except Exception:
        return None
    return None


def classify(files: list[Path]):
    shots = []
    ai = None
    card = None
    others = []
    for f in files:
        ext = f.suffix.lower().lstrip(".")
        name = f.name
        if ext in IMG_EXT:
            shots.append(f)
        elif ext in XLS_EXT:
            kind = task_table_kind(f)
            if kind == "ai" and ai is None:
                ai = f
            elif kind == "card" and card is None:
                card = f
            elif kind is not None:
                continue
            elif kind is None and ai is None and "AI_待分配" in name:
                ai = f
            elif kind is None and card is None and "答题卡_待分配" in name:
                card = f
            else:
                others.append(f)

    valid_others = [x for x in others if has_required_headers(x)]
    if ai is None and valid_others:
        ai = valid_others.pop(0)
    if card is None and valid_others:
        card = valid_others.pop(0)
    return shots, ai, card


def mock_download(ai_target: Path, card_target: Path) -> tuple[bool, bool]:
    ai_src = os.environ.get("DAILY_ASSIGN_MOCK_AI_SOURCE", "").strip()
    card_src = os.environ.get("DAILY_ASSIGN_MOCK_CARD_SOURCE", "").strip()
    ok_ai = False
    ok_card = False
    if ai_src:
        ai_s = Path(ai_src)
        if ai_s.exists():
            shutil.copy2(ai_s, ai_target)
            ok_ai = True
    if card_src:
        card_s = Path(card_src)
        if card_s.exists():
            shutil.copy2(card_s, card_target)
            ok_card = True
    return ok_ai, ok_card


def _start_focus_command(args, wait_seconds=0.08):
    try:
        proc = subprocess.Popen(
            args,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        if wait_seconds <= 0:
            return True
        try:
            return proc.wait(timeout=wait_seconds) == 0
        except subprocess.TimeoutExpired:
            return True
    except Exception:
        return False


def focus_toolbox_app():
    app_path = os.environ.get("TOOLBOX_APP_PATH", "").strip()
    script = """
on run argv
set targetPath to ""
if (count of argv) > 0 then set targetPath to item 1 of argv
try
  if targetPath is not "" then
    set appAlias to POSIX file targetPath as alias
    tell application appAlias to activate
  else
    tell application id "local.liu.Toolbox" to activate
  end if
on error errMsg
  try
    tell application id "local.liu.Toolbox" to activate
  on error
    tell application "Toolbox" to activate
  end try
end try
try
  tell application "System Events"
    set frontmost of first application process whose bundle identifier is "local.liu.Toolbox" to true
  end tell
end try
end run
"""
    requested = False
    if app_path and os.path.exists(app_path):
        requested = _start_focus_command(["/usr/bin/open", app_path]) or requested
    requested = _start_focus_command(["/usr/bin/open", "-b", "local.liu.Toolbox"]) or requested
    _start_focus_command(["/usr/bin/osascript", "-e", script, app_path], wait_seconds=0)
    if requested:
        return


def close_browser_context_and_focus(browser_context):
    focus_toolbox_app()
    try:
        browser_context.close()
    finally:
        focus_toolbox_app()


def normalize_task_text(text: str) -> str:
    if not text:
        return ""
    t = str(text).strip()
    t = re.sub(r"\s+", "", t)
    t = t.replace("（", "(").replace("）", ")")
    return t.lower()


def _query_visible_task_names(page) -> List[str]:
    try:
        data = page.evaluate(
            """() => {
              const out = [];
              // Check menu items
              const menu = Array.from(document.querySelectorAll('ul.ant-menu h5, ul.ant-menu .ant-typography'));
              for (const n of menu) {
                const t = (n?.innerText || '').trim();
                if (t) out.push(t);
              }
              // Check all table cells
              const cells = Array.from(document.querySelectorAll('.ant-table-tbody tr td'));
              for (const n of cells) {
                const t = (n?.innerText || '').trim();
                if (t) out.push(t);
              }
              // Check all links (especially encoded ones in URLs)
              const links = Array.from(document.querySelectorAll('.ant-table-tbody tr a'));
              for (const n of links) {
                try {
                    const href = n?.href || '';
                    if (href) {
                        out.push(href);
                        out.push(decodeURIComponent(href));
                    }
                } catch(e) {}
              }
              return out.slice(0, 200);
            }"""
        )
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _take_screenshot(page, name: str, announce: bool = True):
    try:
        download_dir = Path(os.environ.get("DOWNLOAD_DIR", str(Path.home() / "Downloads")))
        path = download_dir / f"{name}_{int(time.time())}.png"
        path.parent.mkdir(parents=True, exist_ok=True)
        page.screenshot(path=str(path))
        if announce:
            print(f"[截图] 已保存: {path}")
    except Exception as e:
        print(f"[截图] 失败: {e}")


SUBJECT_SUFFIX_RE = re.compile(r"(小学|初中|高中)(语文|数学|英语|物理|化学|生物|科学|历史|地理|政治|道德与法治)$")


def has_subject_suffix(task_name: str) -> bool:
    normalized = normalize_task_text(task_name).rstrip(")）]")
    return bool(SUBJECT_SUFFIX_RE.search(normalized))


def _wait_search_match(page, task_name: str, timeout_ms: int = 6000, require_subject_suffix: bool = False) -> tuple[bool, List[str]]:
    expect = normalize_task_text(task_name)
    elapsed = 0
    sample: List[str] = []
    while elapsed <= timeout_ms:
        rows = _query_visible_task_names(page)
        if rows:
            sample = rows[:8]
            if any(
                expect and expect in normalize_task_text(r) and (not require_subject_suffix or has_subject_suffix(r))
                for r in rows
            ):
                return True, sample
        page.wait_for_timeout(500)
        elapsed += 500
    return False, sample


def _best_subject_task_name_from_page(page, keyword: str = "Auto") -> str | None:
    expect = normalize_task_text(keyword)
    for row in _query_visible_task_names(page):
        text = str(row).strip()
        if expect and expect not in normalize_task_text(text):
            continue
        if has_subject_suffix(text):
            return text
    return None


def _matched_task_names(rows: List[str], task_name: str, require_subject_suffix: bool = False) -> List[str]:
    expect = normalize_task_text(task_name)
    return [
        str(row).strip()
        for row in rows
        if expect
        and expect in normalize_task_text(row)
        and (not require_subject_suffix or has_subject_suffix(row))
    ]


def _visible_button_by_text(page, wanted: str):
    locs = page.locator("button")
    for i in range(min(locs.count(), 80)):
        try:
            btn = locs.nth(i)
            txt = re.sub(r"\s+", "", btn.inner_text(timeout=300) or "")
            if txt == wanted and btn.is_visible() and btn.is_enabled():
                return btn
        except Exception:
            continue
    return None


def _close_open_modals(page):
    try:
        page.evaluate(
            """() => {
              for (const btn of document.querySelectorAll('.ant-modal-close')) {
                const modal = btn.closest('.ant-modal');
                if (modal && getComputedStyle(modal).display !== 'none') btn.click();
              }
            }"""
        )
    except Exception:
        pass


def _click_export_button(page) -> bool:
    try:
        btn = _visible_button_by_text(page, "导出")
        if btn is None:
            return False
        btn.scroll_into_view_if_needed()
        btn.click(timeout=8000, force=True)
        return True
    except Exception:
        return False


def _set_input_value(page, selector: str, value: str, timeout: int = 20000) -> bool:
    try:
        page.wait_for_selector(selector, timeout=timeout)
        page.fill(selector, "")
        page.type(selector, value, delay=8)
        return True
    except Exception:
        return False


def _set_create_time_range(page, start_time: str, end_time: str) -> bool:
    try:
        inputs = page.locator(".ant-picker-range input")
        inputs.first.wait_for(state="visible", timeout=8000)
        page.evaluate(
            """() => {
                document.querySelectorAll('.ant-picker-range input').forEach((el) => {
                    el.removeAttribute('readonly');
                });
            }"""
        )
        inputs.nth(0).click(force=True)
        inputs.nth(0).fill(start_time, force=True)
        page.keyboard.press("Tab")
        inputs.nth(1).fill(end_time, force=True)
        page.keyboard.press("Enter")
        page.keyboard.press("Escape")
        page.wait_for_timeout(250)
        values = inputs.evaluate_all("(els) => els.map((el) => el.value)")
        return len(values) >= 2 and values[0] == start_time and values[1] == end_time
    except Exception:
        return False


def _apply_search_conditions(page, task_name: str, label: str, create_time_range: tuple[str, str] | None = None) -> bool:
    if not _set_input_value(page, 'input#name, input[placeholder*="任务名"]', task_name):
        print(f"E002：无法在 {label} 页面输入任务名")
        _take_screenshot(page, f"{label}_input_fail")
        return False
    if create_time_range is not None:
        start_time, end_time = create_time_range
        if not _set_create_time_range(page, start_time, end_time):
            print(f"E002：无法在 {label} 页面输入创建时间")
            _take_screenshot(page, f"{label}_time_fail")
            return False
    return True


def _trigger_search(page):
    try:
        search_btn = _visible_button_by_text(page, "搜索")
        if search_btn is not None:
            search_btn.click()
        else:
            page.keyboard.press("Enter")

        page.wait_for_timeout(300)
        try:
            loading = page.locator('.ant-spin-spinning, .ant-loading, .ant-table-loading')
            if loading.count() > 0:
                loading.first.wait_for(state="hidden", timeout=6000)
        except Exception:
            pass
    except Exception:
        pass


def _wait_before_export(page, task_name: str, label: str, require_subject_suffix: bool = False) -> None:
    start = time.monotonic()
    last_signature = ""
    stable_count = 0
    min_delay_ms = 3000
    timeout_ms = 9000
    while (time.monotonic() - start) * 1000 <= timeout_ms:
        rows = _query_visible_task_names(page)
        matches = _matched_task_names(rows, task_name, require_subject_suffix)
        signature = "|".join(matches[:8])
        if signature and signature == last_signature:
            stable_count += 1
        else:
            stable_count = 0
            last_signature = signature

        elapsed_ms = (time.monotonic() - start) * 1000
        if elapsed_ms >= min_delay_ms and stable_count >= 1:
            return
        page.wait_for_timeout(500)

    print(f"[下载] {label} 搜索结果已出现，等待导出控件稳定后继续")


def _download_one_table(
    page,
    url: str,
    task_name: str,
    target: Path,
    label: str,
    create_time_range: tuple[str, str] | None = None,
    require_subject_suffix: bool = False,
) -> bool:
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(1200)
    except Exception as e:
        print(f"E002：{label} 页面打开失败 ({e})")
        return False
    
    page.bring_to_front()
    _close_open_modals(page)

    attempted: list[str] = []
    current_task_name = task_name
    for attempt in range(2 if require_subject_suffix else 1):
        attempted.append(current_task_name)
        if not _apply_search_conditions(page, current_task_name, label, create_time_range):
            return False

        _trigger_search(page)

        ok_filter, _ = _wait_search_match(
            page,
            current_task_name,
            timeout_ms=6000,
            require_subject_suffix=require_subject_suffix,
        )
        if not ok_filter:
            print(f"E002：{label} 搜索结果匹配失败（期望：{current_task_name}）")
            _take_screenshot(page, f"{label}_search_fail", announce=False)
            return False

        _wait_before_export(page, current_task_name, label, require_subject_suffix)

        for export_attempt in range(2):
            if not _click_export_button(page):
                print(f"E002：{label} 点击导出按钮失败")
                _take_screenshot(page, f"{label}_export_fail")
                return False

            try:
                modal = page.locator(".ant-modal:visible").last
                modal.wait_for(state="visible", timeout=5000)
                page.wait_for_timeout(120)

                with page.expect_download(timeout=60000) as download_info:
                    confirm = modal.locator('.ant-modal-footer button.ant-btn-primary').last
                    if confirm.count() > 0 and confirm.is_visible(timeout=1200):
                        confirm.click(force=True, timeout=5000)
                    else:
                        page.keyboard.press("Enter")

                download = download_info.value
                tmp_target = target.with_suffix(target.suffix + ".download")
                if tmp_target.exists():
                    tmp_target.unlink()
                download.save_as(str(tmp_target))
                if not tmp_target.exists() or tmp_target.stat().st_size < 500:
                    print(f"E002：{label} 下载文件不完整")
                    return False
                if target.exists():
                    target.unlink()
                tmp_target.replace(target)

                if validate_downloaded_task(target, current_task_name, require_subject_suffix=require_subject_suffix):
                    print(f"[下载] {label} 任务下载成功 ✅")
                    return True

                if require_subject_suffix and attempt == 0:
                    corrected = _best_subject_task_name_from_page(page, task_name)
                    if corrected and corrected not in attempted:
                        print(f"[下载] {label} 下载内容缺少学科后缀，改用任务名关键词重新搜索: {corrected}")
                        current_task_name = corrected
                        break

                if export_attempt == 0:
                    print(f"[下载] {label} 导出内容未匹配当前筛选，重新搜索后再试一次")
                    _close_open_modals(page)
                    _trigger_search(page)
                    _wait_before_export(page, current_task_name, label, require_subject_suffix)
                    continue

                print(f"E002：{label} 下载内容与目标任务不符（可能下载了全部数据或缓存数据）")
                return False
            except Exception as e:
                print(f"E002：{label} 下载确认超时或失败 ({e})")
                _take_screenshot(page, f"{label}_final_fail")
                return False

        if require_subject_suffix and attempted[-1] != current_task_name:
            continue

    return False


def real_download(ai_target: Path, card_target: Path, ai_task: str, card_task: str) -> tuple[bool, bool]:
    user_data_dir = os.path.expanduser("~/.gemini/NewApp_chrome_profile")
    chrome_path = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
    if sync_playwright is None:
        print("E002：Playwright 未安装。")
        return False, False
    if not os.path.exists(chrome_path):
        print(f"E002：找不到 Chrome 路径: {chrome_path}")
        return False, False
    if not os.path.exists(user_data_dir): os.makedirs(user_data_dir, exist_ok=True)

    try:
        with sync_playwright() as p:
            browser_context = p.chromium.launch_persistent_context(
                user_data_dir=user_data_dir, executable_path=chrome_path,
                headless=False, no_viewport=True, accept_downloads=True,
                args=["--remote-debugging-port=9222", "--start-maximized"]
            )
            page = browser_context.pages[0] if browser_context.pages else browser_context.new_page()

            ok1 = _download_one_table(
                page,
                "https://mapi.yuanfudao.com/evaluation/#/admin/evaluation/holepage",
                ai_task,
                ai_target,
                "AI",
                create_time_range=task_create_time_range(date.today()),
                require_subject_suffix=True,
            )
            ok2 = _download_one_table(
                page,
                "https://mapi.yuanfudao.com/evaluation/#/admin/evaluation/card",
                card_task,
                card_target,
                "答题卡",
                create_time_range=task_create_time_range(date.today()),
            )
            close_browser_context_and_focus(browser_context)
            return ok1, ok2
    except Exception as e:
        print(f"E002：自动下载过程异常 ({e})")
        return False, False


def to_num(v) -> float:
    try:
        if v is None: return 0.0
        return float(v)
    except: return 0.0


def downloaded_task_names(path: Path) -> list[str]:
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        header_row, task_col = None, None
        wanted_headers = {"任务名称", "任务名"}
        for r in range(1, min(20, ws.max_row) + 1):
            for c in range(1, min(20, ws.max_column) + 1):
                if str(ws.cell(r, c).value or "").strip() in wanted_headers:
                    header_row, task_col = r, c
                    break
            if header_row is not None:
                break
        if header_row is None or task_col is None:
            return []
        names = []
        for r in range(header_row + 1, ws.max_row + 1):
            v = str(ws.cell(r, task_col).value or "").strip()
            if v:
                names.append(v)
        return names
    except Exception:
        return []


def validate_downloaded_task(path: Path, expected_task_name: str, require_subject_suffix: bool = False) -> bool:
    try:
        names = downloaded_task_names(path)
        if not names:
            return False
        expect = normalize_task_text(expected_task_name)
        matched = [n for n in names if expect and expect in normalize_task_text(n)]
        if not matched:
            return False
        if require_subject_suffix and not any(has_subject_suffix(n) for n in matched):
            return False
        return True
    except Exception:
        return False


def read_tasks(path: Path, cap: int) -> Tuple[List[str], List[TaskRow], float]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    header_row, start_col = None, 3
    header_idx, source_headers = {}, []
    for r in range(1, min(50, ws.max_row) + 1):
        found = False
        for c_start in [3, 1, 2, 4, 5]:
            vals = [str(ws.cell(r, c).value or "").strip() for c in range(c_start, c_start + 11)]
            if REQ_HEADERS.issubset(set(vals)):
                header_row, start_col, source_headers = r, c_start, vals
                for i, v in enumerate(vals, start=c_start):
                    if v: header_idx[v] = i
                found = True
                break
        if found: break
    if header_row is None:
        header_row, start_col = 1, 3
        source_headers = [str(ws.cell(1, c).value or "").strip() for c in range(3, 14)]
        for i, v in enumerate(source_headers, start=3):
            if v: header_idx[v] = i
    out: List[TaskRow] = []
    order = 0
    idx_id = header_idx.get("子任务ID", start_col + 1)
    idx_pages = header_idx.get("总页数", start_col + 7)
    idx_tags = header_idx.get("总评测数量", start_col + 8)
    for r in range(header_row + 1, ws.max_row + 1):
        subtask = str(ws.cell(r, idx_id).value or "").strip()
        pages = to_num(ws.cell(r, idx_pages).value)
        tags = to_num(ws.cell(r, idx_tags).value)
        if not subtask or (pages <= 0 and tags <= 0): continue
        cols = [ws.cell(r, c).value for c in range(start_col, start_col + 11)]
        out.append(TaskRow(order=order, cols=cols, subtask_id=subtask, pages=pages, tags=tags))
        order += 1
    source_pages = sum(page_weight(row) for row in out)
    if cap <= 0:
        return source_headers, [], source_pages
    # DP subset near cap; allow slight overflow to find better balance
    weights = [max(1, int(round(r.pages))) for r in out]
    # Allow more overflow in search to find more precise target
    limit = max(1, int(round(cap * 1.05)))
    prev_sum = [-1] * (limit + 1)
    prev_idx = [-1] * (limit + 1)
    prev_sum[0] = -2
    for i, w in enumerate(weights):
        for s in range(limit, w - 1, -1):
            if prev_sum[s] == -1 and prev_sum[s - w] != -1:
                prev_sum[s] = s - w
                prev_idx[s] = i

    target = int(round(cap))
    best_s = 0
    best_key = (10**9, 10**9)
    for s in range(limit + 1):
        if prev_sum[s] == -1:
            continue
        # Favor small overflow if it's closer to target
        diff = abs(s - target)
        # If error < 2%, it's already quite good
        key = (diff, max(0, s - target)) 
        if key < best_key:
            best_key = key
            best_s = s

    picked_idx = set()
    s = best_s
    while s > 0 and prev_idx[s] != -1:
        i = prev_idx[s]
        picked_idx.add(i)
        s = prev_sum[s]
    picked = [out[i] for i in sorted(picked_idx, key=lambda x: out[x].order)]
    return source_headers, picked, source_pages


def lev(a: str, b: str) -> int:
    if a == b: return 0
    if not a: return len(b)
    if not b: return len(a)
    dp = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        prev = dp[0]; dp[0] = i
        for j, cb in enumerate(b, 1):
            cur = dp[j]
            dp[j] = min(dp[j] + 1, dp[j - 1] + 1, prev + (0 if ca == cb else 1))
            prev = cur
    return dp[-1]


def normalize_name(raw: str, names: List[str]) -> str | None:
    # 1. Basic cleaning
    raw = raw.replace(" ", "").strip()
    if not raw: return None
    if raw in names: return raw
    
    # 2. Hardcode common OCR misreadings
    ocr_map = {"符手娜": "符于娜", "刘兩菲": "刘雨菲", "阎思宇": "阎思宇", "杢梦尻": "李梦园"}
    if raw in ocr_map:
        mapped = ocr_map[raw]
        if mapped in names: return mapped

    # 3. Clean symbols and normalize
    raw_clean = "".join(re.findall(r"[\u4e00-\u9fa5]", raw))
    if "兩" in raw_clean: raw_clean = raw_clean.replace("兩", "雨")
    if "手" in raw_clean and len(raw_clean) == 3 and raw_clean.startswith("符"): 
        raw_clean = raw_clean.replace("手", "于")

    target_clean, spec_target = "韩正", "韩@“”正"
    if raw_clean == target_clean:
        if spec_target in names: return spec_target
        if target_clean in names: return target_clean
        
    # 4. Levenshtein edit distance
    cands = []
    for n in names:
        n_clean = n.replace("@“”", "")
        d = lev(raw_clean, n_clean)
        if d <= 1: cands.append((d, n))
    cands.sort(key=lambda x: x[0])
    if cands:
        if len(cands) == 1 or cands[1][0] > cands[0][0]: return cands[0][1]
    return None


def run_vision_ocr(image_path: Path) -> str:
    bin_env = os.environ.get("OCR_VISION_BIN", "").strip()
    script_env = os.environ.get("OCR_VISION_SCRIPT", "").strip()
    ocr_bin = Path(bin_env) if bin_env else Path(__file__).with_name("ocr_vision_bin")
    script = Path(script_env) if script_env else Path(__file__).with_name("ocr_vision.swift")
    try:
        env = os.environ.copy()
        commands = []
        if ocr_bin.exists() and os.access(ocr_bin, os.X_OK):
            commands.append([str(ocr_bin), str(image_path)])
        if script.exists():
            cache_dir = env.get("TOOLBOX_SWIFT_MODULE_CACHE", "/private/tmp/toolbox_swift_module_cache")
            Path(cache_dir).mkdir(parents=True, exist_ok=True)
            env.setdefault("CLANG_MODULE_CACHE_PATH", cache_dir)
            env.setdefault("SWIFT_MODULE_CACHE_PATH", cache_dir)
            commands.append(["/usr/bin/swift", "-module-cache-path", cache_dir, str(script), str(image_path)])
        for cmd in commands:
            p = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=30,
                env=env,
            )
            if p.returncode == 0 and p.stdout.strip():
                return p.stdout
        return ""
    except: return ""


def load_signup_texts(shots: List[Path]) -> List[str]:
    texts: List[str] = [""] * len(shots)
    missing: List[tuple[int, Path]] = []
    for idx, shot in enumerate(shots):
        text_file = shot.with_suffix(".txt")
        if text_file.exists():
            texts[idx] = text_file.read_text(encoding="utf-8", errors="ignore")
        else:
            missing.append((idx, shot))

    if not missing:
        return texts

    worker_env = os.environ.get("OCR_WORKERS", "").strip()
    try:
        worker_count = int(worker_env) if worker_env else min(3, len(missing))
    except Exception:
        worker_count = min(3, len(missing))
    worker_count = max(1, min(worker_count, len(missing), 6))

    done = 0
    if len(missing) > 1:
        print(f"[识别] 共 {len(missing)} 张截图需要OCR，并行处理 {worker_count} 张...", flush=True)
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {executor.submit(run_vision_ocr, shot): (idx, shot) for idx, shot in missing}
        for future in as_completed(futures):
            idx, _ = futures[future]
            try:
                texts[idx] = future.result()
            except Exception:
                texts[idx] = ""
            done += 1
    return texts


def parse_signup_from_shots(shots: List[Path], names: List[str]) -> Tuple[Dict[str, int], List[str], List[str]]:
    out: Dict[str, int] = {}
    latest_pos: Dict[str, int] = {}
    unmatched: List[str] = []
    unmatched_pos: Dict[str, int] = {}
    noise = {"所有人", "完成", "周五", "周一", "周二", "周三", "周四", "周六", "周日", "评测", "比例", "请大家", "今天", "任务", "表格", "下载", "报名", "自动", "👉", "人数", "约有", "需在", "评测组"}

    def remember(name: str, count: int, pos: int):
        if count not in (2, 3, 5):
            return
        latest_pos[name] = pos
        out[name] = count
        if name in unmatched:
            unmatched.remove(name)
            unmatched_pos.pop(name, None)

    def is_recall_context(text: str, start: int, end: int) -> bool:
        after = text[end:min(len(text), end + 12)]
        return "撤回" in after and "消息" in after

    def parse_count_near(text: str, start: int, end: int) -> int | None:
        left = max(0, start - 8)
        right = min(len(text), end + 12)
        after = text[end:right]
        m = re.search(r"(?<!\d)([235])(?!\d)", after)
        if m:
            return int(m.group(1))
        before = text[left:start]
        nums = [int(x) for x in re.findall(r"(?<!\d)([235])(?!\d)", before)]
        if nums:
            return nums[-1]
        return None

    shot_texts = load_signup_texts(shots)
    for shot_index, text in enumerate(shot_texts):
        compact = re.sub(r"\s+", "", text)
        shot_offset = shot_index * 1_000_000

        for name in names:
            visible_name = name.replace("@“”", "")
            patterns = {name, visible_name}
            if visible_name == "符于娜":
                patterns.update({"符手娜", "符干娜", "符千娜"})
            if visible_name == "刘雨菲":
                patterns.update({"刘兩菲", "刘兩霏", "刘雨霏"})
            for pattern in patterns:
                if not pattern:
                    continue
                for m in re.finditer(re.escape(pattern), compact):
                    if is_recall_context(compact, m.start(), m.end()):
                        continue
                    cnt = parse_count_near(compact, m.start(), m.end())
                    if cnt:
                        remember(name, cnt, shot_offset + m.start())
                        break
                if name in out:
                    break

        for match in re.finditer(r"([\u4e00-\u9fa5@\u201c\u201d\u2018\u2019\u0022\u0027\s]{2,12})[^\d\u4e00-\u9fa5]{0,10}(\d+)", text):
            raw_name, cnt_s = match.group(1), match.group(2)
            if is_recall_context(text, match.start(1), match.end(1)):
                continue
            norm = normalize_name(raw_name, names)
            if not norm:
                pure = "".join(re.findall(r"[\u4e00-\u9fa5]", raw_name))
                if len(pure) < 2 or len(pure) > 4 or any(n in pure for n in noise): continue
                if not any(lev(pure, n.replace("@“”", "")) <= 1 for n in names): continue
                if pure not in unmatched and pure not in out:
                    unmatched.append(pure)
                    unmatched_pos[pure] = shot_offset + match.start()
                continue
            cnt = int(cnt_s)
            if cnt > 0:
                remember(norm, cnt, shot_offset + match.start())

    order = sorted(out.keys(), key=lambda name: (latest_pos.get(name, 10**12), name))
    unmatched.sort(key=lambda name: (unmatched_pos.get(name, 10**12), name))
    return out, unmatched, order


def parse_confirmed_signup(raw: str, names: List[str]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    if not raw.strip():
        return out
    for part in raw.split("|"):
        item = part.strip()
        if not item or ":" not in item:
            continue
        n, c = item.split(":", 1)
        name = n.strip()
        if not name or name not in names:
            continue
        try:
            cnt = int(c.strip())
        except Exception:
            continue
        if cnt not in (2, 3, 5):
            continue
        out[name] = cnt
    return out


def ordered_signup_names(signup: Dict[str, int], name_order: List[str] | None = None) -> List[str]:
    names: List[str] = []
    seen = set()
    for name in name_order or []:
        if name in signup and name not in seen:
            names.append(name)
            seen.add(name)
    for name in signup:
        if name not in seen:
            names.append(name)
            seen.add(name)
    return names


def assign(tasks: List[TaskRow], signup: Dict[str, int], weight_key: str, carry: Dict[str, float] | None = None, name_order: List[str] | None = None):
    names = ordered_signup_names(signup, name_order)
    total_signup = sum(signup.values())
    target_ratio = {n: signup[n] / total_signup for n in names}
    assigned_weight = {n: 0.0 for n in names}
    assigned_count = {n: 0 for n in names}
    if carry:
        for n in names: assigned_weight[n] += carry.get(n, 0.0)
    result, total_w = [], 0.0
    tasks_sorted = sorted(tasks, key=lambda t: (t.pages if weight_key == "page" else t.tags, -t.order), reverse=True)
    for t in tasks_sorted:
        w = t.pages if weight_key == "page" else t.tags
        if w <= 0: w = 1.0
        total_w += w
        best, best_key = None, None
        for n in names:
            debt = target_ratio[n] * max(total_w, 1.0) - assigned_weight[n]
            key = (debt, -assigned_count[n], n)
            if best is None or key > best_key: best, best_key = n, key
        assigned_weight[best] += w
        assigned_count[best] += 1
        result.append([t, best])
    owner_tasks = {n: [] for n in names}
    for i, (_, owner) in enumerate(result):
        owner_tasks[owner].append(i)

    def score() -> tuple[float, float]:
        total = max(sum(assigned_weight.values()), 1.0)
        # Use relative error to target ratio for better precision on different weights
        devs = []
        for n in names:
            target = target_ratio[n]
            if target > 0:
                devs.append(abs(assigned_weight[n] / total - target) / target)
        if not devs: return 0.0, 0.0
        return max(devs), sum(d * d for d in devs)

    cur_score = score()
    # Increase iterations and breadth for better convergence to < 2% relative error
    for iter_idx in range(8000):
        improved = False
        total = max(sum(assigned_weight.values()), 1.0)
        
        # Sort by relative deviation
        signed = []
        for n in names:
            target = target_ratio[n]
            if target > 0:
                signed.append((assigned_weight[n] / total / target - 1.0, n))
        signed.sort()
        
        under = [n for d, n in signed if d < 0]
        over = [n for d, n in reversed(signed) if d > 0]
        
        # Target relative error threshold 2% (0.02)
        if cur_score[0] <= 0.02: 
            break

        # Move: check more candidates
        for o in over[:8]:
            for u in under[:8]:
                best_i = None
                best_local = cur_score
                # Look deeper into task list
                for i in owner_tasks[o][:200]:
                    task, _ = result[i]
                    w = task.pages if weight_key == "page" else task.tags
                    assigned_weight[o] -= w
                    assigned_weight[u] += w
                    new_score = score()
                    assigned_weight[o] += w
                    assigned_weight[u] -= w
                    if new_score < best_local:
                        best_local = new_score
                        best_i = i
                if best_i is not None:
                    task, _ = result[best_i]
                    w = task.pages if weight_key == "page" else task.tags
                    assigned_weight[o] -= w
                    assigned_weight[u] += w
                    result[best_i][1] = u
                    owner_tasks[o].remove(best_i)
                    owner_tasks[u].append(best_i)
                    cur_score = best_local
                    improved = True
                    break
            if improved: break
        if improved: continue

        # Swap: check more pairs
        for o in over[:6]:
            for u in under[:6]:
                best_pair = None
                best_local = cur_score
                for i in owner_tasks[o][:100]:
                    t1, _ = result[i]
                    w1 = t1.pages if weight_key == "page" else t1.tags
                    for j in owner_tasks[u][:100]:
                        t2, _ = result[j]
                        w2 = t2.pages if weight_key == "page" else t2.tags
                        if abs(w1 - w2) < 0.001: continue
                        assigned_weight[o] += (w2 - w1)
                        assigned_weight[u] += (w1 - w2)
                        new_score = score()
                        assigned_weight[o] += (w1 - w2)
                        assigned_weight[u] += (w2 - w1)
                        if new_score < best_local:
                            best_local = new_score
                            best_pair = (i, j, w1, w2)
                if best_pair is not None:
                    i, j, w1, w2 = best_pair
                    result[i][1], result[j][1] = result[j][1], result[i][1]
                    assigned_weight[o] += (w2 - w1)
                    assigned_weight[u] += (w1 - w2)
                    owner_tasks[o].remove(i); owner_tasks[o].append(j)
                    owner_tasks[u].remove(j); owner_tasks[u].append(i)
                    cur_score = best_local
                    improved = True
                    break
            if improved: break
        if not improved: break
    grouped = {n: [] for n in names}
    for item in result: grouped[item[1]].append(tuple(item))
    final = []
    for n in names: final.extend(grouped[n])
    return final, assigned_weight


def task_weight_sum(tasks: List[TaskRow], weight_key: str) -> float:
    total = 0.0
    for t in tasks:
        w = t.pages if weight_key == "page" else t.tags
        total += w if w > 0 else 1.0
    return total


def page_weight(task: TaskRow) -> float:
    return task.pages if task.pages > 0 else 1.0


def normalized_owner_name(owner: str) -> str:
    return str(owner or "").replace("韩@“”正", "韩正").strip()


def make_check_groups(sheet: str, rows) -> List[CheckGroup]:
    grouped: Dict[str, Dict] = {}
    for idx, (task, owner) in enumerate(rows):
        subtask = str(task.subtask_id or "").strip()
        key = subtask if subtask else f"__row_{idx}"
        if key not in grouped:
            grouped[key] = {
                "subtask_id": subtask,
                "row_indices": [],
                "owners": set(),
                "pages": 0.0,
                "first_row_index": idx,
            }
        record = grouped[key]
        record["row_indices"].append(idx)
        owner_name = normalized_owner_name(owner)
        if owner_name:
            record["owners"].add(owner_name)
        record["pages"] += page_weight(task)

    groups: List[CheckGroup] = []
    for key, record in grouped.items():
        groups.append(CheckGroup(
            sheet=sheet,
            key=key,
            subtask_id=record["subtask_id"],
            row_indices=tuple(record["row_indices"]),
            owners=tuple(sorted(record["owners"])),
            pages=record["pages"],
            first_row_index=record["first_row_index"],
        ))
    return sorted(groups, key=lambda g: g.first_row_index)


def sheet_page_total(rows) -> float:
    return sum(page_weight(task) for task, _ in rows)


def random_source() -> random.Random:
    seed = os.environ.get("DAILY_ASSIGN_CHECK_SEED", "").strip()
    if seed:
        return random.Random(seed)
    return random.Random()


def allocate_sheet_checks(
    groups: List[CheckGroup],
    rng: random.Random,
    totals: Dict[str, float],
) -> Dict[tuple[str, int], str]:
    groups = sorted(groups, key=lambda g: g.first_row_index)
    if not groups:
        return {}

    viable = [g for g in groups if any(checker not in g.owners for checker in CHECKERS)]
    if not viable:
        return {}

    target_pages = max(1.0, sum(g.pages for g in groups) * 0.10)
    base_totals = {name: totals.get(name, 0.0) for name in CHECKERS}

    def balance_values(local_totals: Dict[str, float]) -> tuple[float, float]:
        combined = [base_totals.get(name, 0.0) + local_totals.get(name, 0.0) for name in CHECKERS]
        if not combined:
            return 0.0, 0.0
        avg = sum(combined) / len(combined)
        return max(combined) - min(combined), sum((v - avg) ** 2 for v in combined)

    def assign_selected_groups(selected: List[CheckGroup]) -> tuple[Dict[tuple[str, str], str], Dict[str, float]]:
        group_assignments: Dict[tuple[str, str], str] = {}
        local_totals = {name: 0.0 for name in CHECKERS}
        selected = sorted(selected, key=lambda g: (g.pages, -g.first_row_index), reverse=True)

        for group in selected:
            options = [checker for checker in CHECKERS if checker not in group.owners]
            if not options:
                continue
            checker = min(
                options,
                key=lambda name: (base_totals.get(name, 0.0) + local_totals.get(name, 0.0), name),
            )
            group_assignments[(group.sheet, group.key)] = checker
            local_totals[checker] += group.pages

        group_by_key = {(group.sheet, group.key): group for group in selected}

        improved = True
        while improved:
            improved = False
            current = balance_values(local_totals)
            best_move = None
            best_score = current

            for key, checker in list(group_assignments.items()):
                group = group_by_key.get(key)
                if group is None:
                    continue
                for target_checker in CHECKERS:
                    if target_checker == checker or target_checker in group.owners:
                        continue
                    trial = dict(local_totals)
                    trial[checker] -= group.pages
                    trial[target_checker] += group.pages
                    score = balance_values(trial)
                    if score < best_score:
                        best_score = score
                        best_move = ("move", key, checker, target_checker, group.pages)

            keys = list(group_assignments.keys())
            for i, left_key in enumerate(keys):
                left_group = group_by_key.get(left_key)
                left_checker = group_assignments.get(left_key)
                if left_group is None or left_checker is None:
                    continue
                for right_key in keys[i + 1:]:
                    right_group = group_by_key.get(right_key)
                    right_checker = group_assignments.get(right_key)
                    if right_group is None or right_checker is None or left_checker == right_checker:
                        continue
                    if right_checker in left_group.owners or left_checker in right_group.owners:
                        continue
                    trial = dict(local_totals)
                    trial[left_checker] += right_group.pages - left_group.pages
                    trial[right_checker] += left_group.pages - right_group.pages
                    score = balance_values(trial)
                    if score < best_score:
                        best_score = score
                        best_move = ("swap", left_key, right_key, left_checker, right_checker)

            if best_move is None:
                break
            if best_move[0] == "move":
                _, key, old_checker, new_checker, pages = best_move
                group_assignments[key] = new_checker
                local_totals[old_checker] -= pages
                local_totals[new_checker] += pages
            else:
                _, left_key, right_key, left_checker, right_checker = best_move
                left_group = group_by_key[left_key]
                right_group = group_by_key[right_key]
                group_assignments[left_key] = right_checker
                group_assignments[right_key] = left_checker
                local_totals[left_checker] += right_group.pages - left_group.pages
                local_totals[right_checker] += left_group.pages - right_group.pages
            improved = True

        return group_assignments, local_totals

    def build_sample(pool: List[CheckGroup]) -> List[CheckGroup]:
        selected: List[CheckGroup] = []
        pages = 0.0
        for group in pool:
            if not selected or pages < target_pages:
                selected.append(group)
                pages += group.pages
            if pages >= target_pages:
                break
        return selected

    trials = min(1200, max(240, len(viable) * 10))
    trial_pools = [
        sorted(viable, key=lambda g: g.first_row_index),
        sorted(viable, key=lambda g: g.pages, reverse=True),
    ]
    for _ in range(trials):
        pool = viable[:]
        rng.shuffle(pool)
        trial_pools.append(pool)

    best_assignments: Dict[tuple[str, str], str] = {}
    best_totals = {name: 0.0 for name in CHECKERS}
    best_key = (float("inf"), float("inf"), float("inf"), float("inf"), float("inf"))

    for pool in trial_pools:
        selected = build_sample(pool)
        group_assignments, local_totals = assign_selected_groups(selected)
        assigned_pages = sum(local_totals.values())
        if assigned_pages <= 0:
            continue
        balance_range, balance_squares = balance_values(local_totals)
        target_diff = abs(assigned_pages - target_pages)
        under_target = max(0.0, target_pages - assigned_pages)
        key = (target_diff, under_target, balance_range, balance_squares, len(group_assignments))
        if key < best_key:
            best_key = key
            best_assignments = group_assignments
            best_totals = local_totals

    group_by_key = {(group.sheet, group.key): group for group in groups}
    assignments: Dict[tuple[str, int], str] = {}
    for key, checker in best_assignments.items():
        group = group_by_key.get(key)
        if group is None:
            continue
        for row_index in group.row_indices:
            assignments[(group.sheet, row_index)] = checker

    for name, pages in best_totals.items():
        totals[name] = totals.get(name, 0.0) + pages

    return assignments


def plan_checker_assignments(ai_rows, card_rows, mode: str) -> tuple[Dict[tuple[str, int], str], Dict[str, float], Dict[str, float]]:
    rng = random_source()
    all_assignments: Dict[tuple[str, int], str] = {}
    totals = {name: 0.0 for name in CHECKERS}

    ai_groups = make_check_groups("AI", ai_rows)
    card_groups = make_check_groups("答题卡", card_rows)
    sheet_totals = {
        "AI": sheet_page_total(ai_rows),
        "答题卡": sheet_page_total(card_rows),
    }

    if mode == "linked":
        all_assignments.update(allocate_sheet_checks(ai_groups, rng, totals))
        all_assignments.update(allocate_sheet_checks(card_groups, rng, totals))
    else:
        for groups in (ai_groups, card_groups):
            sheet_checker_totals = {name: 0.0 for name in CHECKERS}
            all_assignments.update(allocate_sheet_checks(groups, rng, sheet_checker_totals))
            for name, pages in sheet_checker_totals.items():
                totals[name] += pages

    return all_assignments, sheet_totals, totals


def fmt_pages(value: float) -> str:
    if abs(value - round(value)) < 0.001:
        return str(int(round(value)))
    return f"{value:.1f}".rstrip("0").rstrip(".")


def display_file_link(path: Path | None) -> str | None:
    if path is None:
        return None
    try:
        home = Path.home()
        resolved = path.expanduser()
        downloads = home / "Downloads"
        desktop = home / "Desktop"
        if resolved.parent == downloads:
            return f"Download/{resolved.name}"
        if resolved.parent == desktop:
            return f"Desktop/{resolved.name}"
        return str(resolved.relative_to(home))
    except Exception:
        return str(path)


def print_assignment_ratio_summary(signup: Dict[str, int], signup_order: List[str], rows, method: str) -> None:
    print("[分配] 实际分配数量、比例：")
    total_signup_all = sum(signup.values())
    assigned_w: Dict[str, float] = {}
    for task, owner in rows:
        w = task.pages if method == "page" else task.tags
        assigned_w[owner] = assigned_w.get(owner, 0.0) + (w if w > 0 else 1.0)
    total_weight_all = sum(assigned_w.values())

    display_order = []
    seen = set()
    for name in signup_order:
        if name in signup and name not in seen:
            display_order.append(name)
            seen.add(name)
    for name in signup:
        if name not in seen:
            display_order.append(name)
            seen.add(name)

    for name in display_order:
        signup_n = signup.get(name, 0)
        if signup_n == 0:
            continue
        signup_ratio = signup_n / max(total_signup_all, 1) * 100.0
        assigned = assigned_w.get(name, 0.0)
        assigned_ratio = assigned / max(total_weight_all, 1.0) * 100.0
        warn = " ⚠️" if assigned <= 0 else ""
        print(f"- {name}: 报名 {signup_n}/{signup_ratio:.1f}% | 分配 {fmt_pages(assigned)}/{assigned_ratio:.1f}%{warn}")


def split_signup_for_linked_mode(ai_tasks: List[TaskRow], card_tasks: List[TaskRow], signup: Dict[str, int], weight_key: str, name_order: List[str] | None = None) -> tuple[Dict[str, int], Dict[str, int]]:
    names = ordered_signup_names(signup, name_order)
    if len(names) <= 1 or not ai_tasks or not card_tasks:
        return signup, signup

    ai_total = task_weight_sum(ai_tasks, weight_key)
    card_total = task_weight_sum(card_tasks, weight_key)
    if ai_total <= 0 or card_total <= 0:
        return signup, signup

    total_signup = sum(signup.values())
    target_ai_signup = total_signup * ai_total / (ai_total + card_total)

    best_subset: set[str] | None = None
    best_key = (float("inf"), float("inf"), float("inf"))
    n = len(names)
    # The signup list is small in practice; exhaustive search gives a stable owner partition.
    for mask in range(1, (1 << n) - 1):
        subset = {names[i] for i in range(n) if mask & (1 << i)}
        subset_signup = sum(signup[name] for name in subset)
        key = (
            abs(subset_signup - target_ai_signup),
            abs(len(subset) - n * ai_total / (ai_total + card_total)),
            len(subset),
        )
        if key < best_key:
            best_key = key
            best_subset = subset

    if not best_subset:
        return signup, signup

    ai_signup = {name: signup[name] for name in names if name in best_subset}
    card_signup = {name: signup[name] for name in names if name not in best_subset}
    return ai_signup, card_signup


def ai_task_name(today: date) -> str:
    return "Auto"


def task_create_time_range(today: date) -> tuple[str, str]:
    return f"{today:%Y-%m-%d} 10:00:00", f"{today:%Y-%m-%d} 19:00:00"


def card_task_name(today: date) -> str:
    return "答题卡周期评测"


def delivery_text(today: date) -> str:
    tomorrow = today + timedelta(days=1)
    if today.weekday() == 4: tomorrow = today + timedelta(days=3)
    text = f"{tomorrow.month}月{tomorrow.day}日"
    if tomorrow.weekday() == 3: text += " 14:00"
    return text


def period_text(today: date) -> str:
    if today.day >= 25:
        start = today.replace(day=25)
        end = date(today.year + (1 if today.month == 12 else 0), (today.month % 12) + 1, 24)
    else:
        start = date(today.year - (1 if today.month == 1 else 0), 12 if today.month == 1 else today.month - 1, 25)
        end = today.replace(day=24)
    return f"{start.month:02d}{start.day:02d}-{end.month:02d}{end.day:02d}"


def write_sheet(ws, source_headers: List[str], rows, today: date, checker_assignments: Dict[tuple[str, int], str] | None = None):
    headers = ["周期"] + source_headers + ["负责人", "交付日期", "完成情况", "特殊备注", "检查人", "问责", "是否已修改", "报名截图"]
    ws.append(headers)
    yellow = PatternFill(fill_type="solid", fgColor="FFF4B084")
    for c in "EJK": ws[f"{c}1"].fill = yellow
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}1"
    period, delivery = period_text(today), delivery_text(today)
    sheet_name = ws.title
    checker_assignments = checker_assignments or {}
    for row_index, (task, owner) in enumerate(rows):
        excel_owner = owner.replace("韩@“”正", "韩正")
        checker = checker_assignments.get((sheet_name, row_index), "")
        ws.append([period] + task.cols + [excel_owner, delivery, "", "", checker, "", "", ""])


def main() -> int:
    try: sys.stdout.reconfigure(encoding='utf-8')
    except: pass
    files = split_files(os.environ.get("DAILY_ASSIGN_FILES", ""))
    shots, ai, card = classify(files)
    uploaded_source_files = ai is not None or card is not None
    if not shots:
        print("E001：未检测到 有效 报名截图"); return 1

    names_env = os.environ.get("NAMES", "").strip()
    names = [x.strip() for x in names_env.split(",") if x.strip()]
    signup_ocr, unmatched, ocr_order = parse_signup_from_shots(shots, names)
    preview_rows = [{"name": n, "count": int(signup_ocr.get(n, 0)), "matched": True} for n in ocr_order if signup_ocr.get(n, 0) > 0]
    preview_rows.extend([{"name": n, "count": 5, "matched": False} for n in unmatched])
    print("__OCR_PREVIEW__" + json.dumps({"rows": preview_rows}, ensure_ascii=False))

    if os.environ.get("DAILY_ASSIGN_PREVIEW_ONLY", "0").strip() == "1":
        if not preview_rows:
            print("E004：OCR识别 - 无有效报名数量")
            return 1
        print(f"[识别] 已识别 {len(preview_rows)} 人，请在上方查看/调整后 继续")
        print("（点 金银花，可以返回，重新OCR）")
        return 0

    confirmed_signup = parse_confirmed_signup(os.environ.get("DAILY_ASSIGN_CONFIRMED_SIGNUP", ""), names)
    signup = confirmed_signup if confirmed_signup else signup_ocr
    signup_order = list(confirmed_signup.keys()) if confirmed_signup else list(ocr_order)
    if sum(signup.values()) <= 0:
        print("E004：OCR识别 - 无有效报名数量"); return 1
    print(f"[确认] 报名结果已确认：{len(signup)} 人")

    dl_dir = Path(os.environ.get("DOWNLOAD_DIR", str(Path.home() / "Downloads")))
    dl_dir.mkdir(parents=True, exist_ok=True)
    if ai is None and card is None:
        print("[下载] 检测到仅上传截图，开始自动下载今日任务表...")
        ai_start, ai_end = task_create_time_range(date.today())
        print(f"[下载] AI任务名: {ai_task_name(date.today())}")
        print(f"[下载] AI创建时间: {ai_start}、{ai_end}")
        print(f"[下载] 答题卡任务名: {card_task_name(date.today())}")
        print(f"[下载] 答题卡创建时间: {ai_start}、{ai_end}")
        ai_path, card_path = dl_dir / "AI_待分配.xlsx", dl_dir / "答题卡_待分配.xlsx"
        mode = os.environ.get("DAILY_ASSIGN_DOWNLOAD_MODE", "disabled").strip().lower()
        ok_ai, ok_card = False, False
        if mode == "mock":
            ok_ai, ok_card = mock_download(ai_path, card_path)
        elif mode == "real":
            ok_ai, ok_card = real_download(ai_path, card_path, ai_task_name(date.today()), card_task_name(date.today()))
        if not ok_ai:
            print("\n👉 今天没有AI or 下载失败")
        if not ok_card:
            print("\n👉 今天没有答题卡 or 下载失败")
        ai = ai_path if ok_ai else None
        card = card_path if ok_card else None
    
    method, mode = os.environ.get("DAILY_ASSIGN_METHOD", "page"), os.environ.get("DAILY_ASSIGN_MODE", "independent")
    mode_title = "AI+答题卡" if mode == "linked" else "AI、答题卡"
    print(f"\n[分配] 开始分配 [{mode_title}] ...")

    ai_cap, card_cap = int(os.environ.get("DAILY_ASSIGN_AI_MAX", "200")), int(os.environ.get("DAILY_ASSIGN_CARD_MAX", "300"))
    ai_headers = ["任务名称", "子任务顺序", "任务ID", "子任务ID", "线上学生作业ID", "老师作业ID", "题单ID", "未测评页数", "总页数", "总评测数量", "任务链接"]
    card_headers = ai_headers[:]
    ai_rows, card_rows = [], []
    ai_tasks, card_tasks = [], []
    ai_source_pages, card_source_pages = 0.0, 0.0
    if ai: ai_headers, ai_tasks, ai_source_pages = read_tasks(ai, ai_cap)
    if card: card_headers, card_tasks, card_source_pages = read_tasks(card, card_cap)
    if not ai_tasks and not card_tasks:
        print("E002：没有可分配的 AI 或 答题卡任务表")
        return 1
    if mode == "linked":
        ai_signup, card_signup = split_signup_for_linked_mode(ai_tasks, card_tasks, signup, method, signup_order)
        if ai_tasks: ai_rows, _ = assign(ai_tasks, ai_signup, method, name_order=signup_order)
        if card_tasks: card_rows, _ = assign(card_tasks, card_signup, method, name_order=signup_order)
    else:
        if ai_tasks: ai_rows, _ = assign(ai_tasks, signup, method, name_order=signup_order)
        if card_tasks: card_rows, _ = assign(card_tasks, signup, method, name_order=signup_order)
    checker_assignments, sheet_page_totals, checker_totals = plan_checker_assignments(ai_rows, card_rows, mode)
    output_dir = Path(os.environ.get("OUTPUT_DIR", str(Path.cwd())))
    output, tmp = output_dir / "分配表.xlsx", (output_dir / "分配表.xlsx").with_suffix(".tmp.xlsx")
    try:
        wb = Workbook(); wb.remove(wb.active)
        if ai_rows: write_sheet(wb.create_sheet("AI"), ai_headers, ai_rows, date.today(), checker_assignments)
        if card_rows: write_sheet(wb.create_sheet("答题卡"), card_headers, card_rows, date.today(), checker_assignments)
        wb.save(tmp); os.replace(tmp, output)
    except Exception as e:
        print(f"E005：写出 分配表 失败 ({e})"); return 1
    print_assignment_ratio_summary(signup, signup_order, ai_rows + card_rows, method)
    for bad in sorted(set(unmatched)): print(f"- {bad}: 报名 无法匹配 ⚠️")
    generated_path = display_file_link(output) or f"{output_dir.name}/{output.name}"
    downloaded = [display_file_link(ai), display_file_link(card)]
    downloaded = [p for p in downloaded if p]
    downloaded_line = " , ".join(downloaded)
    source_line_label = "源文件" if uploaded_source_files else "已下载"
    checker_line = "，".join(f"{name} {fmt_pages(checker_totals.get(name, 0.0))} 页" for name in CHECKERS)
    assigned_total_pages = sum(sheet_page_totals.values())
    checked_total_pages = sum(checker_totals.values())
    check_ratio = checked_total_pages / assigned_total_pages * 100.0 if assigned_total_pages > 0 else 0.0
    print("\n------------------------------")
    print(f"[统计] 已分配：AI {fmt_pages(sheet_page_totals.get('AI', 0.0))} 页（共{fmt_pages(ai_source_pages)}页）， 答题卡 {fmt_pages(sheet_page_totals.get('答题卡', 0.0))} 页（共{fmt_pages(card_source_pages)}页）")
    print(f"[分配检查] {checker_line}（共{check_ratio:.1f}%）")
    print("------------------------------")
    if downloaded_line:
        print(f"👉 {source_line_label}：{downloaded_line}\n")
    print(f"👉 已生成：{generated_path}\n")
    print("👉 任务已完成")
    return 0

if __name__ == "__main__": sys.exit(main())
