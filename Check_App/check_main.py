import sys

def _is_help_argv() -> bool:
    return "-h" in sys.argv or "--help" in sys.argv

if _is_help_argv():
    print("""usage: check_main.py [-h] [--base-files BASE_FILES [BASE_FILES ...]]
                     --download-dir DOWNLOAD_DIR --output-file OUTPUT_FILE

options:
  -h, --help            show this help message and exit
  --base-files BASE_FILES [BASE_FILES ...]
  --download-dir DOWNLOAD_DIR
  --output-file OUTPUT_FILE""")
    sys.exit(0)

# 立即打印启动信息（--help 时不打印，避免干扰 argparse）
if not _is_help_argv():
    print(" - 正在扫描上传文件...")
    print("=" * 50)
    sys.stdout.flush()

import os
import argparse
import unicodedata
import time
import subprocess
from collections import defaultdict
from pathlib import Path

# 延迟导入重量级库
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    sync_playwright = None  # type: ignore

# ── 样式常量 ──────────────────────────────────────────────────────────────
FONT = Font(name="等线", size=12)
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
GREEN_FILL  = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

def display_width(s):
    """准确计算终端显示宽度"""
    w = 0
    for char in str(s):
        if unicodedata.east_asian_width(char) in ('W', 'F', 'A'):
            w += 2
        else:
            w += 1
    return w

def pad_display(s, width):
    curr_w = display_width(s)
    return s + " " * max(0, width - curr_w)

def log_status(label, status, width=42):
    print(f"{pad_display(label, width)}{status}")

DEBUG_TIMING = os.environ.get("TOOLBOX_WEEKLY_DEBUG_TIMING") == "1"

def debug_timing(start_time, label):
    if DEBUG_TIMING:
        print(f"[耗时] {label}: {time.perf_counter() - start_time:.2f}s")
        sys.stdout.flush()


def focus_toolbox_app():
    app_path = os.environ.get("TOOLBOX_APP_PATH", "").strip()
    script = """
on run argv
set targetPath to ""
if (count of argv) > 0 then set targetPath to item 1 of argv
try
  if targetPath is not "" then
    set appAlias to POSIX file targetPath as alias
    tell application appAlias to activate
  else
    tell application id "local.liu.Toolbox" to activate
  end if
on error errMsg
  try
    tell application id "local.liu.Toolbox" to activate
  on error
    tell application "Toolbox" to activate
  end try
end try
try
  tell application "System Events"
    set frontmost of first application process whose bundle identifier is "local.liu.Toolbox" to true
  end tell
end try
end run
"""
    for delay in (0, 0.2, 0.6):
        if delay:
            time.sleep(delay)
        try:
            if app_path and os.path.exists(app_path):
                subprocess.run(
                    ["/usr/bin/open", app_path],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    timeout=3,
                    check=False,
                )
            subprocess.run(
                ["/usr/bin/osascript", "-e", script, app_path],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=3,
                check=False,
            )
        except Exception:
            pass

# ── 工具函数 ─────────────────────────────────────────────────────────────
def read_sheet(ws):
    it = ws.iter_rows(values_only=True)
    try:
        first_row = next(it)
    except StopIteration:
        return [], []
    headers = [str(c) if c is not None else "" for c in first_row]
    data = [list(r) for r in it]
    return headers, data

def col(headers, name):
    lower_name = name.lower()
    for i, h in enumerate(headers):
        if h.lower() == lower_name:
            return i
    return -1

def dedup_by(rows, headers, col_name):
    idx = col(headers, col_name)
    if idx < 0: return rows
    seen, out = set(), []
    for r in rows:
        if idx >= len(r): continue
        key = r[idx]
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out

def convert_link(link, page_type="holepage"):
    if not link: return ""
    s = str(link)
    return s.replace(f"#/admin/evaluation/{page_type}/", f"#/evaluation/{page_type}/")\
            .replace("#/admin/evaluation/holepage/", "#/evaluation/holepage/")\
            .replace("#/admin/evaluation/cardPage/", "#/evaluation/cardPage/")

def apply_format(ws, yellow_cols=None, skip_width_col=None):
    yellow_cols = yellow_cols or []
    headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    for h_name in yellow_cols:
        for i, h in enumerate(headers, 1):
            if h == h_name:
                ws.cell(1, i).fill = YELLOW_FILL
    
    max_sample = min(100, ws.max_row)
    for ci in range(1, ws.max_column + 1):
        col_letter = get_column_letter(ci)
        if not (skip_width_col and ci == skip_width_col):
            max_w = 0
            for ri in range(1, max_sample + 1):
                v = ws.cell(ri, ci).value
                if v:
                    max_w = max(max_w, display_width(str(v)))
            if max_w:
                ws.column_dimensions[col_letter].width = min(max_w + 2, 60)
        for ri in range(1, ws.max_row + 1):
            cell = ws.cell(ri, ci)
            if cell.value is not None: cell.font = FONT

DOWNLOAD_HEADER_CHECKERS = {
    "分数.xlsx": lambda h: "题目框识别情况" in h and "是否有分数框" in h and "大题ID" in h,
    "AI.xlsx": lambda h: "阶段" in h and "科目" in h and "任务名" in h,
    "答题卡-分数.xlsx": lambda h: "答案框识别情况" in h and "分数框识别情况" in h and "题目框识别情况" not in h,
    "答题卡-AI.xlsx": lambda h: "任务名" in h and "任务id" in h and "手写识别" in h and "阶段" not in h,
}

def first_row_header_set(ws):
    try:
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return set()
    return {str(c) if c is not None else "" for c in row}

def downloaded_file_matches(path, expected_name):
    checker = DOWNLOAD_HEADER_CHECKERS.get(expected_name)
    if checker is None:
        return True
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            for sname in wb.sheetnames:
                if checker(first_row_header_set(wb[sname])):
                    return True
        finally:
            wb.close()
    except Exception:
        return False
    return False

# ── 数据扫描 ──────────────────────────────────────────────────────────────
def scan_download_dir(download_dir):
    NAMED_HEADERS = dict(DOWNLOAD_HEADER_CHECKERS)
    named_data = {} 

    if not os.path.exists(download_dir): return named_data

    files = []
    for fname in os.listdir(download_dir):
        if not fname.endswith(".xlsx") or fname.startswith("~"): continue
        fpath = os.path.join(download_dir, fname)
        try:
            files.append((fpath, fname, os.path.getmtime(fpath)))
        except: pass
        
    files.sort(key=lambda x: x[2], reverse=True)

    for fpath, fname, _ in files[:50]:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                headers, rows = read_sheet(ws)
                if not headers: continue
                h_set = set(headers)

                for norm_name, checker in list(NAMED_HEADERS.items()):
                    if checker(h_set):
                        named_data[norm_name] = (fname, headers, rows)
                        del NAMED_HEADERS[norm_name]
                        break
            wb.close()
            if not NAMED_HEADERS:
                break
        except: continue
    return named_data

def build_owner_map(base_info):
    owner = {}
    for headers, rows, _ in base_info:
        zi_idx, fu_idx = col(headers, "子任务ID"), col(headers, "负责人")
        if zi_idx < 0 or fu_idx < 0: continue
        for r in rows:
            zi = r[zi_idx] if zi_idx < len(r) else None
            fu = r[fu_idx] if fu_idx < len(r) else None
            if zi is not None: owner[str(zi).lower()] = str(fu or "")
    return owner

def lookup_owner(row, headers, owner_map):
    zi_idx = col(headers, "子任务id")
    if zi_idx < 0 or zi_idx >= len(row): return ""
    return owner_map.get(str(row[zi_idx]).lower(), "") if row[zi_idx] is not None else ""

# ── Sheet 生成 ────────────────────────────────────────────────────────────
def build_link_sheet(wb, base_info):
    ws = wb.create_sheet("拼链接")
    ws.append(["名称", "链接"])
    s2_ids, s1_ids = set(), set()
    for h, rows, is_card in base_info:
        idx = col(h, "任务ID")
        if idx < 0: continue
        target_set = s1_ids if is_card else s2_ids
        for r in rows:
            if idx < len(r) and r[idx] is not None: target_set.add(str(r[idx]))
    
    url1 = "https://mapi.yuanfudao.com/evaluation/#/evaluation/holepage/statistics?taskIds=[" + ",".join(s2_ids) + "]" if s2_ids else ""
    url2 = "https://mapi.yuanfudao.com/evaluation/#/evaluation/cardPage/statistics?taskIds=[" + ",".join(s1_ids) + "]" if s1_ids else ""
    ws.append(["分数+固定_全链接", url1])
    ws.append(["答题卡_链接", url2])
    apply_format(ws, skip_width_col=2)
    return url1, url2

def generic_build(wb, named_data, sheet_name, file_key, filter_fn, owner_map, link_page="holepage", yellow_cols=None):
    if file_key not in named_data: return 0
    ws = wb.create_sheet(sheet_name)
    fname, h, rows = named_data[file_key]
    link_idx = col(h, "链接")
    filtered = [r for r in rows if filter_fn(r, h)]
    filtered = dedup_by(filtered, h, "链接")
    ws.append(h + ["负责人", "兼职用的链接"])
    for r in filtered:
        lnk = r[link_idx] if link_idx >= 0 and link_idx < len(r) else ""
        ws.append(r + [lookup_owner(r, h, owner_map), convert_link(lnk, link_page)])
    apply_format(ws, yellow_cols=yellow_cols)
    return len(filtered)

# ── 自动下载逻辑 ──────────────────────────────────────────────────────────
def auto_download_files(url1, url2, download_dir, label_width):
    user_data_dir = os.path.expanduser("~/.gemini/NewApp_chrome_profile")
    chrome_path = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
    if sync_playwright is None:
        print("\n错误：Playwright 未安装或打包不完整，无法自动下载明细表。请使用源码环境运行或重新打包 check_main_bin。")
        sys.stdout.flush()
        return False
    if not os.path.exists(chrome_path):
        print(f"\n错误：未找到 Google Chrome（期望路径：{chrome_path}）。自动下载需要本机已安装 Chrome。")
        sys.stdout.flush()
        return False
    if not os.path.exists(user_data_dir): os.makedirs(user_data_dir, exist_ok=True)

    print("\n...下载更新 数据源表...")
    sys.stdout.flush()
    
    try:
        with sync_playwright() as p:
            timing_start = time.perf_counter()
            browser_context = p.chromium.launch_persistent_context(
                user_data_dir=user_data_dir, executable_path=chrome_path,
                headless=False, no_viewport=True, accept_downloads=True,
                args=["--remote-debugging-port=9222", "--start-maximized"]
            )
            page = browser_context.pages[0] if browser_context.pages else browser_context.new_page()

            def save_download(download_obj, target_path, expected_name):
                if os.path.exists(target_path):
                    os.remove(target_path)
                download_obj.save_as(target_path)
                if not os.path.exists(target_path) or os.path.getsize(target_path) <= 0:
                    return False
                ok = downloaded_file_matches(target_path, expected_name)
                if not ok:
                    try:
                        os.remove(target_path)
                    except OSError:
                        pass
                return ok

            def first_clickable_button(text):
                locators = [
                    page.get_by_role("button", name=text),
                    page.locator(f"button:has-text('{text}')"),
                ]
                for locator in locators:
                    try:
                        count = min(locator.count(), 4)
                    except Exception:
                        count = 0
                    for idx in range(count):
                        btn = locator.nth(idx)
                        try:
                            if not btn.is_visible(timeout=500):
                                continue
                            btn.scroll_into_view_if_needed(timeout=1000)
                            if btn.is_enabled(timeout=500):
                                return btn
                        except Exception:
                            continue
                return None

            def download_by_button_text(button_text, target_file, total_timeout=60000, attempts=3):
                target_path = os.path.join(download_dir, target_file)
                deadline = time.monotonic() + (total_timeout / 1000.0)
                per_attempt_timeout = max(1000, total_timeout // max(attempts, 1))

                def remaining_ms():
                    return max(0, int((deadline - time.monotonic()) * 1000))

                def wait_for_detail_ready():
                    texts = ["导出评测详情", "题目框答案框评测详细统计信息", "任务名"]
                    if target_file == "AI.xlsx":
                        texts.append("阶段")
                    elif target_file == "答题卡-AI.xlsx":
                        texts.append("手写识别")

                    remaining = remaining_ms()
                    if remaining <= 0:
                        return
                    try:
                        page.wait_for_function(
                            "texts => texts.every(text => (document.body && document.body.innerText || '').includes(text))",
                            arg=texts,
                            timeout=min(20000, remaining),
                        )
                    except Exception:
                        pass

                for attempt in range(attempts):
                    try:
                        wait_for_detail_ready()
                        remaining = remaining_ms()
                        if remaining <= 0:
                            break
                        page.wait_for_timeout(min(500 + attempt * 250, remaining))
                        btn = first_clickable_button(button_text)
                        if btn is None:
                            raise RuntimeError(f"未找到按钮：{button_text}")
                        remaining = remaining_ms()
                        if remaining <= 0:
                            break
                        with page.expect_download(timeout=min(per_attempt_timeout, remaining)) as d:
                            btn.click(force=True, timeout=min(5000, max(1000, remaining)))
                        if save_download(d.value, target_path, target_file):
                            return True
                    except Exception:
                        remaining = remaining_ms()
                        if remaining <= 0:
                            break
                        page.wait_for_timeout(min(1000, remaining))
                return False

            def open_ai_tab_if_present():
                for tab_name in ("AI批改评测统计信息", "AI 批改业务维度统计信息"):
                    try:
                        tabs = page.get_by_role("tab", name=tab_name)
                        if tabs.count() > 0 and tabs.first.is_visible(timeout=500):
                            tabs.first.click(force=True, timeout=3000)
                            page.wait_for_timeout(500)
                            return
                    except Exception:
                        pass

            def process_link(url, is_card=False):
                link_label = "答题卡" if is_card else "AI"
                page.goto(url, wait_until="domcontentloaded", timeout=120000)
                debug_timing(timing_start, f"{link_label} 页面打开")
                try: page.wait_for_url(lambda u: "mapi.yuanfudao.com" in u, timeout=120000)
                except: pass
                page.bring_to_front()
                open_ai_tab_if_present()

                # AI.xlsx
                ai_file = "答题卡-AI.xlsx" if is_card else "AI.xlsx"
                ok = download_by_button_text("导出评测详情", ai_file, total_timeout=60000, attempts=3)
                debug_timing(timing_start, f"{ai_file} 下载")
                log_status(ai_file.replace('-', '_'), "✅" if ok else "❌", label_width)
                sys.stdout.flush()

                # Score.xlsx
                try:
                    tab = page.get_by_role("tab", name="分数识别评测统计信息")
                    tab.click(force=True)
                    try: page.wait_for_selector("text=分数识别每页评测详情统计信息", timeout=10000)
                    except: tab.click(force=True); page.wait_for_timeout(3000)
                    
                    for _ in range(8):
                        if page.locator("button:has-text('导出Excel')").count() >= 2: break
                        page.evaluate("window.scrollBy(0, 1000)"); time.sleep(1)
                    
                    score_file = "答题卡-分数.xlsx" if is_card else "分数.xlsx"
                    btns = page.locator("button:has-text('导出Excel')")
                    if btns.count() >= 2:
                        with page.expect_download(timeout=180000) as d2: btns.last.click(force=True)
                        ok = save_download(d2.value, os.path.join(download_dir, score_file), score_file)
                        debug_timing(timing_start, f"{score_file} 下载")
                        log_status(score_file.replace('-', '_'), "✅" if ok else "❌", label_width)
                    else:
                        log_status(score_file.replace('-', '_'), "❌", label_width)
                except:
                    log_status(('答题卡-分数.xlsx' if is_card else '分数.xlsx').replace('-', '_'), "❌", label_width)
                sys.stdout.flush()

            if url1: process_link(url1, is_card=False)
            if url2: process_link(url2, is_card=True)
            browser_context.close()
            focus_toolbox_app()
            print("")
            return True
    except Exception as e:
        print(f"\n错误：自动下载失败：{e}")
        sys.stdout.flush()
        return False



# ── 主流程 ────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-files", nargs='+', default=[])
    parser.add_argument("--download-dir", required=True)
    parser.add_argument("--output-file", required=True)
    args = parser.parse_args()

    base_files = args.base_files
    download_dir = args.download_dir
    out_path = args.output_file

    base_info = []
    for fpath in base_files:
        if not os.path.exists(fpath): continue
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                headers, rows = read_sheet(ws)
                if not headers: continue
                h_set = set(headers)
                if "负责人" in h_set:
                    b_idx = col(headers, "任务名称")
                    has_datika = any("答题卡" in str(r[b_idx]) for r in rows[:20] if b_idx >=0 and b_idx < len(r) and r[b_idx])
                    base_info.append((headers, rows, has_datika))
            wb.close()
        except: continue

    if not base_info:
        print("错误：未找到有效的基础任务文件")
        return

    LABEL_WIDTH = 42

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    url1, url2 = build_link_sheet(wb, base_info)
    log_status("拼链接 ...", "已生成 ✅", LABEL_WIDTH)
    sys.stdout.flush()

    # 自动下载
    if url1 or url2:
        auto_download_files(url1, url2, download_dir, LABEL_WIDTH)
        
    named = scan_download_dir(download_dir)
    
    owner_map = build_owner_map(base_info)
    
    # 过滤器定义
    f_score = lambda r, h: col(h, "是否有分数框") >= 0 and col(h, "分数识别情况") >= 0 and r[col(h, "是否有分数框")] == "有分数框" and r[col(h, "分数识别情况")] in ("忽略", "错误")
    f_fixed = lambda r, h: col(h, "任务名") >= 0 and col(h, "固定批改识别结果") >= 0 and r[col(h, "任务名")] and "固定" in str(r[col(h, "任务名")]) and r[col(h, "固定批改识别结果")] in ("忽略", "错误")

    # 1. 分数识别
    c = generic_build(wb, named, "分数识别错误或忽略", "分数.xlsx", f_score, owner_map, yellow_cols=["是否有分数框", "分数识别情况"])
    log_status("分数识别错误或忽略 ...", f"已生成 ✅ ({c}行)", LABEL_WIDTH)
    
    # 2. 固定批改
    c = generic_build(wb, named, "固定批改错误或忽略", "AI.xlsx", f_fixed, owner_map, yellow_cols=["固定批改识别结果"])
    log_status("固定批改错误或忽略 ...", f"已生成 ✅ ({c}行)", LABEL_WIDTH)
    
    # 3. 题目框
    if "AI.xlsx" in named:
        ws = wb.create_sheet("题目框错误+忽略超过3个")
        _, h, rows = named["AI.xlsx"]
        m_idx, lk_idx, bid_idx = col(h, "题目框识别情况"), col(h, "链接"), col(h, "大题ID")
        errs = [r for r in rows if m_idx>=0 and r[m_idx]=="错误"]
        igns = [r for r in rows if m_idx>=0 and r[m_idx]=="忽略"]
        lk_map = defaultdict(set)
        for r in igns: lk_map[r[lk_idx]].add(r[bid_idx])
        valid_lk = {lk for lk, bids in lk_map.items() if len(bids) >= 3}
        final = dedup_by(errs + [r for r in igns if r[lk_idx] in valid_lk], h, "链接")
        ws.append(h + ["负责人", "兼职用的链接"])
        for r in final: ws.append(r + [lookup_owner(r, h, owner_map), convert_link(r[lk_idx])])
        apply_format(ws, yellow_cols=["题目框识别情况"])
        log_status("题目框错误+忽略超过3个 ...", f"已生成 ✅ ({len(final)}行)", LABEL_WIDTH)
    else: log_status("题目框错误+忽略超过3个 ...", "等待补充 ⏳", LABEL_WIDTH)

    # 4. BadCase
    if "AI.xlsx" in named:
        ws = wb.create_sheet("BadCase")
        _, h, rows = named["AI.xlsx"]
        p, v, w = col(h, "批改情况"), col(h, "是否AI可批"), col(h, "作答结果-算法可解评测结果")
        flt = [r for r in rows if p>=0 and v>=0 and w>=0 and r[p]=="错误" and r[v]=="AI可批" and r[w]=="是"]
        ws.append(h + ["负责人", "是否修改", "检查人"])
        for r in flt: ws.append(list(r) + [lookup_owner(r, h, owner_map), "", ""])
        apply_format(ws, yellow_cols=["批改情况", "是否AI可批", "作答结果-算法可解评测结果"])
        if len(flt) >= 2: ws.cell(row=int(round(len(flt)/2.0))+1, column=len(h)+3).fill = GREEN_FILL
        log_status("BadCase ...", f"已生成 ✅ ({len(flt)}行)", LABEL_WIDTH)
    else: log_status("BadCase ...", "等待补充 ⏳", LABEL_WIDTH)

    # 5. 答题卡_分数
    f_c_score = lambda r, h: col(h, "是否有分数框")>=0 and col(h, "分数识别情况")>=0 and col(h, "分数-算法可解评测结果")>=0 and r[col(h, "是否有分数框")]=="有分数框" and r[col(h, "分数识别情况")] in ("忽略", "错误") and r[col(h, "分数-算法可解评测结果")]=="是"
    c = generic_build(wb, named, "答题卡_分数识别错误或忽略", "答题卡-分数.xlsx", f_c_score, owner_map, "cardPage", ["是否有分数框", "分数识别情况", "分数-算法可解评测结果"])
    log_status("答题卡_分数识别错误或忽略 ...", f"已生成 ✅ ({c}行)", LABEL_WIDTH)

    # 6. 答题卡_AI_BadCase
    if "答题卡-AI.xlsx" in named:
        ws = wb.create_sheet("答题卡_AI_BadCase")
        _, h, rows = named["答题卡-AI.xlsx"]
        n, p, lk = col(h, "批改情况"), col(h, "作答结果-算法可解评测结果"), col(h, "链接")
        flt = [r for r in rows if n>=0 and p>=0 and r[n]=="错误" and r[p]=="是"]
        ws.append(h + ["负责人", "兼职用的链接"])
        for r in flt: ws.append(r + [lookup_owner(r, h, owner_map), convert_link(r[lk], "cardPage")])
        apply_format(ws, yellow_cols=["批改情况", "作答结果-算法可解评测结果"])
        log_status("答题卡_AI_BadCase ...", f"已生成 ✅ ({len(flt)}行)", LABEL_WIDTH)
    else: log_status("答题卡_AI_BadCase ...", "等待补充 ⏳", LABEL_WIDTH)

    # 7. 答题卡_AI
    if "答题卡-AI.xlsx" in named:
        ws = wb.create_sheet("答题卡_AI_错误次数≥3（星标参考）")
        _, h, rows = named["答题卡-AI.xlsx"]
        n, p, lk, bid = col(h, "批改情况"), col(h, "作答结果-算法可解评测结果"), col(h, "链接"), col(h, "大题ID")
        flt = [r for r in rows if n>=0 and p>=0 and r[n]=="错误" and r[p]=="是"]
        cnts = defaultdict(int)
        for r in flt: cnts[r[bid]] += 1
        vld_b = {b for b, count in cnts.items() if count >= 3}
        fnl = dedup_by([r for r in flt if r[bid] in vld_b], h, "链接")
        
        # 重新排序，大题ID相同的挨在一起
        fnl.sort(key=lambda x: x[bid])
        
        ws.append(h + ["负责人", "兼职用的链接"])
        for r in fnl: ws.append(r + [lookup_owner(r, h, owner_map), convert_link(r[lk], "cardPage")])
        
        # 大题ID交替灰色底色
        GRAY_FILL_S7 = PatternFill(start_color="FFD0D0D0", end_color="FFD0D0D0", fill_type="solid")
        curr_bid, fill_st = None, False
        for ri in range(2, ws.max_row + 1):
            b_val = ws.cell(row=ri, column=bid+1).value
            if b_val != curr_bid:
                curr_bid, fill_st = b_val, not fill_st
            if fill_st:
                ws.cell(row=ri, column=bid+1).fill = GRAY_FILL_S7

        apply_format(ws, yellow_cols=["批改情况", "作答结果-算法可解评测结果"])
        log_status("答题卡_AI_错误次数≥3（星标参考） ...", f"已生成 ✅ ({len(fnl)}行)", LABEL_WIDTH)
    else: log_status("答题卡_AI_错误次数≥3（星标参考） ...", "等待补充 ⏳", LABEL_WIDTH)

    # 8. AI_错误次数≥3（星标参考）
    if "AI.xlsx" in named:
        ws = wb.create_sheet("AI_错误次数≥3（星标参考）")
        _, h, rows = named["AI.xlsx"]
        p, v, w = col(h, "批改情况"), col(h, "是否AI可批"), col(h, "作答结果-算法可解评测结果")
        lk_idx, bid_idx = col(h, "链接"), col(h, "大题ID")
        task_idx = col(h, "任务名")
        
        flt = [r for r in rows if p>=0 and v>=0 and w>=0 and r[p]=="错误" and r[v]=="AI可批" and r[w]=="是"]
        bids_links = defaultdict(set)
        for r in flt: bids_links[r[bid_idx]].add(r[lk_idx])
        valid_bids = {bid for bid, links in bids_links.items() if len(links) >= 3}
        flt2 = [r for r in flt if r[bid_idx] in valid_bids]
        
        # 结果按链接+大题ID去重
        final = []
        seen_pairs = set()
        for r in flt2:
            pair = (r[lk_idx], r[bid_idx])
            if pair not in seen_pairs:
                seen_pairs.add(pair)
                final.append(r)
        
        # 排序：分数在上面，固定在下面；大题ID相同的挨在一起
        final.sort(key=lambda x: (0 if "分数" in str(x[task_idx]) else (1 if "固定" in str(x[task_idx]) else 2), x[bid_idx]))
        
        ws.append(h + ["负责人"])
        GRAY_FILL_L = PatternFill(start_color="FFD0D0D0", end_color="FFD0D0D0", fill_type="solid")
        YELLOW_FILL_L = PatternFill(start_color="FFFFFFE0", end_color="FFFFFFE0", fill_type="solid")
        
        last_score_row_idx = -1
        for i, r in enumerate(final):
            ws.append(list(r) + [lookup_owner(r, h, owner_map)])
            if "分数" in str(r[task_idx]):
                last_score_row_idx = i + 2
        
        # 最后一条“分数”对应的链接，单元格加灰色底色
        if last_score_row_idx != -1:
            ws.cell(row=last_score_row_idx, column=lk_idx+1).fill = GRAY_FILL_L
            
        # H列大题ID交替底色
        curr_bid, fill_st = None, False
        for ri in range(2, ws.max_row + 1):
            b_val = ws.cell(row=ri, column=bid_idx+1).value
            if b_val != curr_bid:
                curr_bid, fill_st = b_val, not fill_st
            if fill_st:
                ws.cell(row=ri, column=bid_idx+1).fill = GRAY_FILL_L
        
        # H列表头加浅黄色底色
        ws.cell(row=1, column=bid_idx+1).fill = YELLOW_FILL_L
        
        # 隐藏列
        hide_indices = [1, 2, 4, 5, 6, 7, 9, 10, 11, 12, 13, 14, 15, 17, 18, 19, 20, 21] + list(range(24, 38))
        for ci in hide_indices:
            ws.column_dimensions[get_column_letter(ci)].hidden = True
            
        apply_format(ws, yellow_cols=["批改情况", "是否AI可批", "作答结果-算法可解评测结果"])
        log_status("AI_错误次数≥3（星标参考） ...", f"已生成 ✅ ({len(final)}行)", LABEL_WIDTH)
    else: log_status("AI_错误次数≥3（星标参考） ...", "等待补充 ⏳", LABEL_WIDTH)

    wb.save(out_path)

    output = Path(out_path)
    folder_name = output.parent.name or "."
    print("=" * 51 + f"\n👉 已生成：{folder_name}/{output.name}\n")

if __name__ == "__main__": main()
